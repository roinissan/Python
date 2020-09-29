from openpyxl.styles import Alignment,colors
from openpyxl import Workbook,load_workbook
from openpyxl.styles.borders import Border,Side
from openpyxl.styles import PatternFill


class StyleExcel:
    def __init__(self,plc_file):
        self.plc_file_path = plc_file

    def centeraize(self,sheet):
        for row in sheet.iter_rows(min_row = sheet.min_row, min_col = sheet.min_column, max_row = sheet.max_row , max_col = sheet.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')


    def column_rows_adjust(self,sheet):
        dims = {}
        for row in sheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
        #try:
        for col, value in dims.items():
            sheet.column_dimensions[col].width = value
        #except Exception:
         #   print(Exception)
          #  print(col)
           # print(value)

    def borders(self,sheet):
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        for row in sheet.iter_rows(min_row = sheet.min_row, min_col = sheet.min_column, max_row = sheet.max_row , max_col = sheet.max_column):
            for cell in row:
                cell.border = thin_border

    def color_ok_reject_cells(self,sheet):
        for row in sheet.iter_rows(min_row = sheet.min_row, min_col = sheet.min_column, max_row = sheet.max_row , max_col = sheet.max_column):
            for cell in row:
                if (cell.value == "OK"):
                   cell.fill = PatternFill('solid',colors.GREEN)
                elif (cell.value == "REJECT"):
                    cell.fill = PatternFill('solid',colors.RED)




    def run(self):
        worksheet = load_workbook(self.plc_file_path)
        sheet = worksheet.active
        self.centeraize(sheet)
        self.column_rows_adjust(sheet)
        self.borders(sheet)
        self.color_ok_reject_cells(sheet)
        worksheet.save(self.plc_file_path)


